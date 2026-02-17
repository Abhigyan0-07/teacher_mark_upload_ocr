import base64
import io
from pathlib import Path
from typing import List

import cv2
import numpy as np
import pytesseract
from PIL import Image
from openpyxl import Workbook, load_workbook
from fastapi import HTTPException


def _decode_base64_image(image_b64: str) -> np.ndarray:
    header, _, data = image_b64.partition(",")
    if data:
        image_b64 = data
    img_bytes = base64.b64decode(image_b64)
    img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    open_cv_image = np.array(img)
    open_cv_image = open_cv_image[:, :, ::-1].copy()  # RGB -> BGR
    return open_cv_image


def _ocr_box(image: np.ndarray) -> int:
    # 1. Grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # 2. Upscale for better contour detection
    scale = 2.0
    width = int(gray.shape[1] * scale)
    height = int(gray.shape[0] * scale)
    dim = (width, height)
    resized = cv2.resize(gray, dim, interpolation=cv2.INTER_CUBIC)

    # 3. Threshold (Otsu) - Invert first to find white text on black background for contours
    # But usually paper is white, text is black. 
    # Thresholding: 
    #   cv2.THRESH_BINARY_INV: Text becomes White, Background Black (Best for finding contours of text)
    _, th_inv = cv2.threshold(resized, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    # 4. Find Contours
    # Use RETR_LIST to find all contours, including those inside a border
    contours, _ = cv2.findContours(th_inv, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

    # 5. Filter Candidates
    # If no contours found, it's empty
    if not contours:
        return 0

    # Filter out very small noise
    valid_contours = [c for c in contours if cv2.contourArea(c) > 20]
    
    # Filter out likely grid lines / borders
    # 1. Touches edge?
    # 2. Long and thin?
    
    final_contours = []
    h_img, w_img = height, width
    
    for c in valid_contours:
        x, y, w, h = cv2.boundingRect(c)
        
        # Check if touches edge (within 3px)
        touches_edge = (x <= 3) or (y <= 3) or (x + w >= w_img - 3) or (y + h >= h_img - 3)
        
        # Check if long (spans > 50% of dimension)
        spans_long = (w > 0.5 * w_img) or (h > 0.5 * h_img)
        
        if touches_edge and spans_long:
            continue # Skip border line
            
        final_contours.append(c)
    
    if not final_contours:
        return 0
        
    # Find bounding box that encompasses ALL valid contours (for multi-digit support)
    min_x, min_y = width, height
    max_x, max_y = 0, 0
    
    # print(f"[DEBUG] Valid contours found: {len(final_contours)}")

    for c in final_contours:
        x, y, w, h = cv2.boundingRect(c)
        min_x = min(min_x, x)
        min_y = min(min_y, y)
        max_x = max(max_x, x + w)
        max_y = max(max_y, y + h)
        
    x, y = min_x, min_y
    w = max_x - min_x
    h = max_y - min_y
    
    # 6. Extract ROI and Center it
    # Let's get a clean black-text-white-bg version of the whole image first
    _, th_clean = cv2.threshold(resized, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # Sanity check
    if w <= 0 or h <= 0:
        return 0
        
    digit_roi = th_clean[y:y+h, x:x+w]
    
    # Create a square canvas to center the content
    # Size should be max(w, h) + padding
    side = max(w, h) + 40 # 20px padding
    canvas = np.ones((side, side), dtype=np.uint8) * 255 # White canvas
    
    # Center coordinates
    c_x = side // 2
    c_y = side // 2
    
    # Top-left dict
    dx = c_x - w // 2
    dy = c_y - h // 2
    
    # Paste ROI
    canvas[dy:dy+h, dx:dx+w] = digit_roi
    
    print(f"[DEBUG] Valid contours found: {len(final_contours)}")

    # ...

    # Save debug image (Removed for production, can re-enable if needed)
    import os
    import time
    debug_dir = r"C:\Users\abhig\OneDrive\Desktop\marks\backend\debug_crops"
    os.makedirs(debug_dir, exist_ok=True)
    timestamp = int(time.time() * 1000)
    cv2.imwrite(f"{debug_dir}/{timestamp}_debug.png", canvas)
    
    # Save raw crop too for comparison if possible? 
    # Let's just save the processed canvas for now as it shows what Tesseract sees.

    # 7. Final OCR
    # ...
    
    try:
        # Define config for Tesseract (e.g., to recognize only digits)
        # Assuming 'config' is defined elsewhere or needs to be added.
        # For this change, we'll assume it's implicitly handled or will be added.
        # If not, pytesseract.image_to_string(canvas) would be used.
        config = r'--oem 3 --psm 8 -c tessedit_char_whitelist=0123456789'
        text = pytesseract.image_to_string(canvas, config=config)
        print(f"[DEBUG] Raw OCR text: '{text.strip()}'")
    except pytesseract.TesseractNotFoundError:
        print("[ERROR] Tesseract not found.")
        raise Exception("Tesseract OCR is not installed on the server. Please install it.")
    except Exception as e:
        print(f"[ERROR] OCR failed: {e}")
        return 0
        
    digits = "".join(ch for ch in text if ch.isdigit())
    val = int(digits) if digits else 0
    print(f"[DEBUG] Parsed value: {val}")
    return val



from ..ocr.google_vision import detect_document_text

def _get_centroids(vertices):
    """
    Returns (x, y) centroid of a polygon defined by vertices.
    vertices: list of objects with .x and .y
    """
    if not vertices:
        return 0, 0
    
    xs = [v.x for v in vertices]
    ys = [v.y for v in vertices]
    return sum(xs) / len(xs), sum(ys) / len(ys)

# Configure logging
import logging
import sys

# Create a custom logger
logger = logging.getLogger("backend_debug")
logger.setLevel(logging.DEBUG)

# Check if handlers already exist to avoid duplicates
if not logger.handlers:
    # File Handler
    fh = logging.FileHandler('backend_debug.log')
    fh.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    logger.addHandler(fh)

    # Stream Handler (stdout)
    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(logging.DEBUG)
    sh.setFormatter(formatter)
    logger.addHandler(sh)

# Use 'logger' instead of 'logging' in the rest of the file
# We'll need to update calls, or just alias logging=logger (risky but quick)
# Better to just use 'logger' global.


def extract_grid_marks(image_b64: str, rows: int = 4, cols: int = 2) -> List[int]:
    """
    Extract marks from a fixed grid using Google Cloud Vision API.
    Sends the WHOLE image to Google Vision once, then maps detected text 
    to the corresponding grid cell based on coordinates.
    """
    image_cv = _decode_base64_image(image_b64)
    h, w = image_cv.shape[:2]
    logger.info(f"Processing image: {w}x{h}, Rows={rows}, Cols={cols}")
    
    # Encode back to bytes for Google Vision
    if "," in image_b64:
        _, _, data = image_b64.partition(",")
        img_bytes = base64.b64decode(data)
    else:
        img_bytes = base64.b64decode(image_b64)
        
    try:
        annotation = detect_document_text(img_bytes)
    except Exception as e:
        logger.error(f"Google Vision API failed: {e}")
        print(f"[ERROR] Google Vision API failed: {e}")
        print("Falling back to legacy local OCR...")
        return _extract_grid_marks_fallback(image_cv, rows, cols)


    # Initialize grid marks
    grid_marks = [0] * (rows * cols)
    
    row_height = h / rows
    col_width = w / cols
    
    found_marks = [] # List of {'val': int, 'x': float, 'y': float} 
    
    # Iterate through all pages/blocks/paragraphs/words/symbols
    for page in annotation.pages:
        for block in page.blocks:
            for paragraph in block.paragraphs:
                for word in paragraph.words:
                    word_text = "".join([symbol.text for symbol in word.symbols])
                    
                    # Clean the text
                    # If text contains non-digits, clean them.
                    cleaned_text = "".join(filter(str.isdigit, word_text))
                    if not cleaned_text:
                        continue
                        
                    val = int(cleaned_text)
                    
                    # Get bounding box
                    vertices = word.bounding_box.vertices
                    cx, cy = _get_centroids(vertices)
                    
                    # Logic: 
                    # If val <= 100, assume it's a single mark.
                    # If val > 100 (e.g. 547, 105), it's likely merged digits.
                    # Exception: 100 is valid.
                    
                    if val <= 100:
                        logger.debug(f"Found Mark: {val} at ({cx}, {cy})")
                        found_marks.append({'val': val, 'x': cx, 'y': cy})
                    else:
                        # Split! "547" -> 5, 4, 7
                        logger.debug(f"Splitting merged value: {val} at ({cx}, {cy})")
                        s_val = str(val)
                        n_chars = len(s_val)
                        
                        # Interpolate X coordinates
                        # Approximate width of the word box
                        min_x = min([v.x for v in vertices])
                        max_x = max([v.x for v in vertices])
                        width = max_x - min_x
                        char_width = width / n_chars
                        
                        for i, char in enumerate(s_val):
                            digit = int(char)
                            # Estimate center of this char
                            char_cx = min_x + (i * char_width) + (char_width / 2)
                            
                            logger.debug(f"  -> Split: {digit} at ({char_cx}, {cy})")
                            found_marks.append({'val': digit, 'x': char_cx, 'y': cy})

    # --- SMART GRID LOGIC ---
    total_cells = rows * cols
    logger.info(f"Total candidates found: {len(found_marks)}. Expected: {total_cells}")
    
    # NEW STRATEGY: Content-Based Grid Inference
    # Instead of assuming the grid fills the image, we infer the grid bounds from the detected Marks.
    if len(found_marks) >= 3: # Need at least a few points to infer a grid
        logger.info("Attempting to infer grid from available points.")
        try:
             return _infer_grid_from_candidates(found_marks, rows, cols)
        except Exception as e:
             logger.error(f"Inferred Grid Mapping failed: {e}")

    logger.warning("Fallback to Basic Rigid Grid (Image-Based).")
    logger.debug(f"[DEBUG] Only {len(found_marks)} marks found vs {total_cells} expected. Using Rigid Fallback.")
    
    # Case B: Standard Rigid Grid Mapping (Fallback)
    grid_marks = [0] * total_cells
    
    for item in found_marks:
        val = item['val']
        cx = item['x']
        cy = item['y']
        
        r = int(cy // row_height)
        c = int(cx // col_width)
        
        logger.debug(f"Mapping {val} at ({cx}, {cy}) to Cell [{r}, {c}]")
        
        if 0 <= r < rows and 0 <= c < cols:
            idx = r * cols + c
            grid_marks[idx] = val

    return grid_marks

def _infer_grid_from_candidates(candidates: List[dict], rows: int, cols: int) -> List[int]:
    """
    Infers the grid structure based on the bounding box of the detected numbers.
    """
    if not candidates:
        return [0] * (rows * cols)

    # 1. Determine Bounding Box of Content
    min_x = min(c['x'] for c in candidates)
    max_x = max(c['x'] for c in candidates)
    min_y = min(c['y'] for c in candidates)
    max_y = max(c['y'] for c in candidates)
    
    width = max_x - min_x
    height = max_y - min_y
    
    # Avoid zero division
    if width < 10: width = 100
    if height < 10: height = 100
    
    # Expand bounds slightly to cover proper cell area
    # (min_x, min_y) is usually the center of the top-left digit.
    # We want top-left of the GRID.
    # Approx cell size
    avg_cell_h = height / max(1, (rows - 1))
    avg_cell_w = width / max(1, (cols - 1))
    
    start_x = min_x - (avg_cell_w / 2)
    start_y = min_y - (avg_cell_h / 2)
    end_x = max_x + (avg_cell_w / 2)
    end_y = max_y + (avg_cell_h / 2)
    
    eff_w = end_x - start_x
    eff_h = end_y - start_y
    
    logger.debug(f"Inferred Grid Bounds: x={start_x:.1f}-{end_x:.1f}, y={start_y:.1f}-{end_y:.1f}")
    
    grid_marks = [0] * (rows * cols)
    
    for item in candidates:
        # Normalize to 0-1 within the effective grid
        rel_x = (item['x'] - start_x) / eff_w
        rel_y = (item['y'] - start_y) / eff_h
        
        # Clamp
        rel_x = max(0, min(1, rel_x))
        rel_y = max(0, min(1, rel_y))
        
        # Map to index
        c = int(rel_x * cols)
        r = int(rel_y * rows)
        
        # Edge case
        if c >= cols: c = cols - 1
        if r >= rows: r = rows - 1
        
        idx = r * cols + c
        
        if grid_marks[idx] == 0:
            grid_marks[idx] = item['val']
        else:
            # Overwrite collision
            grid_marks[idx] = item['val']
            
    return grid_marks

def _smart_grid_cluster(candidates: List[dict], rows: int, cols: int, img_h: int, img_w: int) -> List[int]:
    """
    Robustly maps a list of candidate points {'val', 'x', 'y'} to a grid.
    Handles cases where len(candidates) >= rows * cols (noise, labels).
    Uses 1D clustering on Y then X.
    """
    logger.info(f"Smart Clustering: {len(candidates)} candidates for {rows}x{cols} grid")
    
    # 1. Cluster Rows (Y-axis)
    # Sort by Y
    by_y = sorted(candidates, key=lambda k: k['y'])
    
    # Simple Gap Detection
    # Iterate and split if gap > threshold
    # Threshold heuristic: The image is divided into 'rows'. 
    # A gap between rows should be at least (img_h / rows) * 0.3
    row_gap_thresh = (img_h / rows) * 0.3
    
    current_row = [by_y[0]]
    row_groups = []
    
    for i in range(1, len(by_y)):
        curr = by_y[i]
        prev = by_y[i-1]
        gap = curr['y'] - prev['y']
        
        if gap > row_gap_thresh:
            # New row
            row_groups.append(current_row)
            current_row = [curr]
        else:
            current_row.append(curr)
    row_groups.append(current_row)
    
    logger.info(f"  -> Found {len(row_groups)} vertical clusters (rows). Expected {rows}.")
    
    # Heuristic: If we found more rows than expected, merge small/close ones or take largest?
    # If we found fewer, we might have merged two rows.
    # For now, let's try to map strictly if count matches, otherwise picking best matches.
    
    # Filter Row Groups:
    # If we have > rows, we might have noise at top/bottom.
    # Sort groups by size (number of elements)? No, valid rows might be sparse.
    # Sort groups by average Y and pick the ones that align with expected rigid lines?
    # Let's simple-sort by Centroid Y and take the top 'rows' groups if they look valid?
    # Better: Re-Assign based on rigid expectations if clustering fails?
    # Let's stick to the gaps. If we have > rows, merge the closest ones?
    
    # If we have too many row groups, take the ones that contain the most items?
    if len(row_groups) > rows:
        row_groups.sort(key=lambda g: len(g), reverse=True)
        row_groups = row_groups[:rows]
        # Re-sort by Y
        row_groups.sort(key=lambda g: sum(i['y'] for i in g)/len(g))
        
    final_grid = [0] * (rows * cols)
    
    for r_idx, r_group in enumerate(row_groups):
        if r_idx >= rows: break
        
        # 2. Cluster Cols (X-axis) within this row
        by_x = sorted(r_group, key=lambda k: k['x'])
        
        col_gap_thresh = (img_w / cols) * 0.3
        
        col_groups = []
        curr_col = [by_x[0]]
        
        for i in range(1, len(by_x)):
            curr = by_x[i]
            prev = by_x[i-1]
            gap = curr['x'] - prev['x']
            
            if gap > col_gap_thresh:
                col_groups.append(curr_col)
                curr_col = [curr]
            else:
                curr_col.append(curr)
        col_groups.append(curr_col)
        
        # Taking top 'cols' groups
        if len(col_groups) > cols:
             # Sort by X position to pick "left, middle, right" logic? 
             # No, if we have extra, it might be "Label Value".
             # If "Label Value" (gap small), they are in SAME group.
             # If "Value Value" (gap large), different groups.
             
             # If we have too many distinct column groups, filtering is tricky.
             # Assume taking the ones with best X-alignment?
             pass
             
        # Map to valid columns
        # For each group, determine its column index based on image width
        for c_group in col_groups:
            # Determine representative X
            avg_x = sum(i['x'] for i in c_group) / len(c_group)
            
            # Map avg_x to column index semi-rigidly relative to image width
            # (Just to assign valid slot)
            c_idx = int(avg_x / (img_w / cols))
            if c_idx >= cols: c_idx = cols - 1
            
            # Use the item in the group.
            # If multiple items in one column group (e.g. "Q1" "5" very close),
            # pick the RIGHTMOST one (Marks usually follow labels)
            # OR pick the distinct value <= 100
            
            # Sort by X descending
            c_group.sort(key=lambda k: k['x'], reverse=True)
            chosen_val = c_group[0]['val'] # Rightmost
            
            flat_idx = r_idx * cols + c_idx
            final_grid[flat_idx] = chosen_val
            logger.debug(f"    -> Mapped val {chosen_val} to [{r_idx},{c_idx}]")

    return final_grid


# Regex to find numbers, ignoring common labels
import re

def _clean_and_find_mark(text: str) -> int | None:
    """
    Heuristic to find the most likely mark in a text string.
    1. Ignores "Q1", "No.1", etc.
    2. Looks for standalone numbers.
    3. Returns the valid mark (<= 100) if found.
    """
    # Replace common non-digit chars that might be confused or attached
    # e.g. "5/10" -> "5 10"
    text = text.replace("/", " ").replace("|", " ")
    
    # Find all number-like tokens
    # This finds "1", "10", "Q1" (we'll filter Q1 later)
    tokens = re.findall(r'[a-zA-Z]*\d+', text)
    
    candidates = []
    for token in tokens:
        # If it starts with Q or No, ignore (e.g. Q1, No5)
        if re.match(r'^[qQnN]\d+', token):
            continue
            
        # Clean to just digits
        digits = "".join(filter(str.isdigit, token))
        if not digits:
            continue
            
        val = int(digits)
        # Marks are usually 0-100
        if 0 <= val <= 100:
            candidates.append(val)
            
    if not candidates:
        return None
        
    # If multiple candidates?
    # e.g. "10 (2)" -> [10, 2]. Usually the mark is the main number.
    # e.g. "Q1 5" -> "5" (Q1 filtered).
    # If we have [10, 2], it's ambiguous. But often the mark is the largest or the last one?
    # Or maybe the one that is NOT the question number.
    # Let's take the LAST one, assuming user writes mark at end or right side?
    # Or return the MAX?
    # Let's return the one that looks most like a mark.
    return candidates[-1]

def extract_single_mark(image_b64: str) -> int:
    """
    Extract a single mark from a cropped image.
    Uses Google Cloud Vision API effectively on the small crop.
    """
    if "," in image_b64:
        _, _, data = image_b64.partition(",")
        img_bytes = base64.b64decode(data)
    else:
        img_bytes = base64.b64decode(image_b64)
    
    try:
        annotation = detect_document_text(img_bytes)
        full_text = annotation.text or ""
        print(f"[DEBUG] Manual Crop Text: {full_text}")
        
        val = _clean_and_find_mark(full_text)
        if val is not None:
            return val
        return 0
        
    except Exception as e:
        print(f"[ERROR] Google Vision API failed on single crop: {e}")
        # Fallback to local
        image_cv = _decode_base64_image(image_b64)
        return _ocr_box(image_cv)


from pytesseract import Output

def _extract_grid_marks_fallback(image: np.ndarray, rows: int = 4, cols: int = 2) -> List[int]:
    """
    Legacy Tesseract implementation UPGRADED with Smart Sort.
    Uses image_to_data on full image instead of slicing,
    then applies the same spatial logic.
    """
    h, w = image.shape[:2]
    logger.info(f"[Fallback] Processing image: {w}x{h}, Rows={rows}, Cols={cols}")
    
    # Preprocess for Tesseract
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # Adaptive Threshold (better for shadows/uneven lighting)
    # Block size 31 (larger) to not break thin lines of large digits
    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 2)
    
    # Optional: Morphological opening to remove small noise
    kernel = np.ones((1,1), np.uint8)
    thresh = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel)
    
    # Run Tesseract on FULL image to get boxes
    # PSM 11 = Sparse text. Find as much text as possible in no particular order.
    # OEM 3 = Default (LSTM + Legacy)
    custom_config = r'--oem 3 --psm 11 outputbase digits' 
    
    logger.info(f"[Fallback-Debug] Running Tesseract with config: {custom_config}")
    
    try:
        d = pytesseract.image_to_data(thresh, config=custom_config, output_type=Output.DICT)
    except Exception as e:
        logger.error(f"[ERROR] Tesseract failed: {e}")
        return [0] * (rows * cols)

    found_marks = []
    
    n_boxes = len(d['text'])
    logger.info(f"[Fallback-Debug] Tesseract found {n_boxes} potential text blocks.")
    
    for i in range(n_boxes):
        conf = int(d['conf'][i])
        text = d['text'][i].strip()
        
        # Log EVERYTHING
        # logger.debug(f"[Fallback-Debug] Raw Box {i}: '{text}' (conf={conf})")
        
        if text:
            # Clean
            clean_text = "".join(filter(str.isdigit, text))
            if not clean_text:
                logger.info(f"[Fallback-Debug] Ignored non-digit: '{text}' (conf={conf})")
                continue
                
            val = int(clean_text)
            x, y, w_box, h_box = d['left'][i], d['top'][i], d['width'][i], d['height'][i]
            cx = x + w_box / 2
            cy = y + h_box / 2
            
            logger.info(f"[Fallback-Debug] Found Candidate: {val} (raw='{text}', conf={conf}) at ({cx:.1f}, {cy:.1f})")
            
            # Logic: Split if merged
            if val <= 100:
                found_marks.append({'val': val, 'x': cx, 'y': cy})
            else:
                # Split merged digits (e.g. 547 -> 5, 4, 7)
                logger.info(f"[Fallback-Debug] Splitting merged: {val}")
                s_val = str(val)
                n_chars = len(s_val)
                char_width = w_box / n_chars
                
                for k, char in enumerate(s_val):
                    digit = int(char)
                    char_cx = x + (k * char_width) + (char_width / 2)
                    found_marks.append({'val': digit, 'x': char_cx, 'y': cy})


    # --- SHARED SMART GRID LOGIC ---
    # Case A: Sufficient candidates found
    total_cells = rows * cols
    logger.info(f"[Fallback] Found {len(found_marks)} candidates. Expected: {total_cells}")
    
    if len(found_marks) >= total_cells:
        logger.info("[Fallback] Sufficient candidates found. Using Smart Clustering.")
        try:
             return _smart_grid_cluster(found_marks, rows, cols, h, w)
        except Exception as e:
             logger.error(f"Smart Clustering failed: {e}")
             # Fall through to rigid grid
    
    # Case B: Rigid Fallback
    logger.warning("[Fallback] Mismatch or Clustering Failed. Using Rigid Grid.")
    
    grid_marks = [0] * total_cells
    row_height = h / rows
    col_width = w / cols

    for item in found_marks:
        val = item['val']
        cx = item['x']
        cy = item['y']
        
        r = int(cy // row_height)
        c = int(cx // col_width)
        
        if 0 <= r < rows and 0 <= c < cols:
            idx = r * cols + c
            grid_marks[idx] = val

    return grid_marks

def append_marks_to_excel(
    marks: List[int],
    excel_content: bytes | None = None
) -> tuple[int, bytes]:
    """
    Append a row with marks and total to an Excel sheet.
    If excel_content is None, creates a new workbook.
    Returns (total, modified_excel_bytes).
    """
    total = sum(marks)

    if excel_content:
        # Load from bytes
        wb = load_workbook(io.BytesIO(excel_content))
        ws = wb.active
    else:
        # Create new
        wb = Workbook()
        ws = wb.active
        # Header row
        for idx in range(len(marks)):
            ws.cell(row=1, column=idx + 1).value = f"Q{idx + 1}"
        ws.cell(row=1, column=len(marks) + 1).value = "Total"

    row = ws.max_row + 1
    for idx, value in enumerate(marks):
        ws.cell(row=row, column=idx + 1).value = value
    ws.cell(row=row, column=len(marks) + 1).value = total

    # Save to bytes
    out_buffer = io.BytesIO()
    wb.save(out_buffer)
    out_buffer.seek(0)
    return total, out_buffer.getvalue()
